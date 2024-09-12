import platform
import subprocess
from PIL import Image
from io import BytesIO
import win32clipboard
import openpyxl
import re
import pyperclip
import tkinter as tk
from tkinter import filedialog, ttk
import threading
from tkinter import messagebox
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException, WebDriverException
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.chrome.options import Options
import os
import shutil
from PIL import ImageGrab, Image
from datetime import datetime


class WhatsAppAutomation:

    def __init__(self, root):
        self.root = root
        self.root.title("WhatsApp Automation")
        self.root.geometry("600x500")
        self.root.resizable(True, True)

        self.init_gui()

        self.driver = None
        self.wait = None
        self.filepath = ''
        self.webdriver_path = ''
        self.photo_path = ''
        self.stop_thread = False
        self.pause_thread = False
        self.process_thread = None  # For concurrency management

    def init_gui(self):
        main_frame = tk.Frame(self.root, padx=10, pady=10)
        main_frame.pack(fill="both", expand=True)

        header = tk.Label(main_frame, text="WhatsApp Automation", font=("Arial", 18, "bold"), pady=10)
        header.pack()

        btn_webdriver = tk.Button(main_frame, text="Set WebDriver Path", command=self.set_webdriver_path, padx=10, pady=5)
        btn_webdriver.pack(pady=5)

        btn_login = tk.Button(main_frame, text="Login", command=self.login, padx=10, pady=5)
        btn_login.pack(pady=5)

        btn_choose_file = tk.Button(main_frame, text="Choose Excel File", command=self.choose_file, padx=10, pady=5)
        btn_choose_file.pack(pady=5)

        self.chosen_language = tk.StringVar(value="en")
        self.send_mode = tk.StringVar(value="message")

        lang_frame = tk.Frame(main_frame)
        lang_frame.pack(pady=5)
        lang_en = tk.Radiobutton(lang_frame, text="English", variable=self.chosen_language, value="en")
        lang_ar = tk.Radiobutton(lang_frame, text="Arabic", variable=self.chosen_language, value="ar")
        lang_en.pack(side="left", padx=5)
        lang_ar.pack(side="right", padx=5)

        mode_frame = tk.Frame(main_frame)
        mode_frame.pack(pady=5)
        mode_message = tk.Radiobutton(mode_frame, text="Send Message", variable=self.send_mode, value="message", command=self.toggle_photo_options)
        mode_photo = tk.Radiobutton(mode_frame, text="Send Photo", variable=self.send_mode, value="photo", command=self.toggle_photo_options)
        mode_message.pack(side="left", padx=5)
        mode_photo.pack(side="right", padx=5)

        self.btn_choose_photo = tk.Button(main_frame, text="Choose Photo", command=self.choose_photo, padx=10, pady=5)
        self.btn_choose_photo.pack(pady=5)
        self.btn_choose_photo.pack_forget()

        btn_send = tk.Button(main_frame, text="Send", command=self.start_process, padx=10, pady=5, bg="green", fg="white")
        btn_send.pack(pady=5)

        self.btn_pause = tk.Button(main_frame, text="Pause", command=self.pause_messages, padx=10, pady=5, bg="orange", fg="white")
        self.btn_pause.pack(pady=5)

        btn_stop = tk.Button(main_frame, text="Stop", command=self.stop_messages, padx=10, pady=5, bg="red", fg="white")
        btn_stop.pack(pady=5)

        self.info_var = tk.StringVar()
        self.info_var.set("Status: Waiting...")
        lbl_info = tk.Label(main_frame, textvariable=self.info_var, pady=10, font=("Arial", 12))
        lbl_info.pack()

        log_frame = tk.Frame(main_frame)
        log_frame.pack(fill="both", expand=True)

        self.text_area = tk.Text(log_frame, height=10, width=50, font=("Arial", 10), padx=10, pady=10)
        self.scrollbar = tk.Scrollbar(log_frame, command=self.text_area.yview)
        self.text_area.configure(yscrollcommand=self.scrollbar.set)
        self.text_area.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

        self.progress = ttk.Progressbar(main_frame, orient="horizontal", mode="determinate", length=500)
        self.progress.pack(pady=10)

        self.time_var = tk.StringVar()
        self.time_var.set("Estimated Time Remaining: --:--")
        lbl_time = tk.Label(main_frame, textvariable=self.time_var, pady=10, font=("Arial", 12))
        lbl_time.pack()

    def toggle_photo_options(self):
        if self.send_mode.get() == "photo":
            self.btn_choose_photo.pack(pady=5)
        else:
            self.btn_choose_photo.pack_forget()

    def choose_photo(self):
        self.photo_path = filedialog.askopenfilename(title="Select Photo", filetypes=[("Image files", "*.jpg;*.jpeg;*.png"), ("All files", "*.*")])
        if self.photo_path:
            target_path = os.path.join(os.getcwd(), os.path.basename(self.photo_path))
            shutil.copy(self.photo_path, target_path)
            self.photo_path = target_path
            self.update_text_area(f"Photo chosen: {self.photo_path}")
        else:
            self.update_info_var("No photo chosen.")
            self.update_text_area("No photo chosen.")

    def copy_image_to_clipboard(self, image_path):
        system_platform = platform.system()

        if system_platform == "Windows":
            image = Image.open(image_path)
            output = BytesIO()
            image.convert("RGB").save(output, "BMP")
            data = output.getvalue()[14:]
            output.close()

            win32clipboard.OpenClipboard()
            win32clipboard.EmptyClipboard()
            win32clipboard.SetClipboardData(win32clipboard.CF_DIB, data)
            win32clipboard.CloseClipboard()

        elif system_platform == "Linux":
            image = Image.open(image_path)
            output = BytesIO()
            image.save(output, "PNG")
            output.seek(0)

            process = subprocess.Popen(['xclip', '-selection', 'clipboard', '-t', 'image/png'], stdin=subprocess.PIPE)
            process.communicate(input=output.read())

        elif system_platform == "Darwin":
            image = Image.open(image_path)
            output = BytesIO()
            image.save(output, "PNG")
            output.seek(0)

            process = subprocess.Popen(['osascript', '-e', 'set the clipboard to (read (POSIX file "{}") as JPEG picture)'.format(image_path)],
                                    stdin=subprocess.PIPE)
            process.communicate(input=output.read())

        else:
            raise NotImplementedError(f"Clipboard copy is not implemented for {system_platform}.")

    def start_process(self):
        """Start the message sending process in a separate thread for concurrency management."""
        if not self.process_thread or not self.process_thread.is_alive():
            self.process_thread = threading.Thread(target=self.send_messages)
            self.process_thread.start()

    def pause_messages(self):
        self.pause_thread = not self.pause_thread
        if self.pause_thread:
            self.update_info_var("Paused...")
            self.btn_pause.config(text="Resume", bg="blue")  # Update button to "Resume"
        else:
            self.update_info_var("Resuming...")
            self.btn_pause.config(text="Pause", bg="orange")  # Revert button to "Pause"

    def stop_messages(self):
        """Stop the message sending process."""
        self.stop_thread = True
        self.update_info_var("Stopping...")
        self.update_text_area("Attempting to stop the message sending process...")
        if self.process_thread and self.process_thread.is_alive():
            self.process_thread.join()  # Ensure the thread is terminated correctly
        self.update_info_var("Process Stopped.")

    def set_webdriver_path(self):
        self.webdriver_path = filedialog.askopenfilename(title="Select WebDriver", filetypes=[("Executable files", "*.exe"), ("All files", "*.*")])
        if self.webdriver_path and os.path.isfile(self.webdriver_path):
            self.update_info_var(f"WebDriver set to: {self.webdriver_path}")
            self.update_text_area(f"WebDriver set to: {self.webdriver_path}")
        else:
            self.update_info_var("Invalid WebDriver path!")
            self.update_text_area("Invalid WebDriver path!")

    def login(self):
        if not self.webdriver_path:
            self.update_info_var("Please select WebDriver path first!")
            self.update_text_area("Please select WebDriver path first!")
            return

        chrome_options = Options()
        data_dir = os.path.join(os.getcwd(), 'whatsapp_session')
        chrome_options.add_argument(f"--user-data-dir={data_dir}")

        try:
            service = webdriver.chrome.service.Service(self.webdriver_path)
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            self.wait = WebDriverWait(self.driver, 10)
        except WebDriverException as e:
            self.update_info_var(f"Failed to launch WebDriver: {str(e)}")
            self.update_text_area(f"Error launching WebDriver: {str(e)}")
            return

        try:
            self.driver.get("https://web.whatsapp.com/")
            self.update_info_var("Status: Please login to WhatsApp...")
            self.update_text_area("Opened WhatsApp Web. Please login.")
        except Exception as e:
            self.update_info_var(f"Failed to open WhatsApp Web: {str(e)}")
            self.update_text_area(f"Error opening WhatsApp Web: {str(e)}")

    def choose_file(self):
        self.filepath = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel files", "*.xlsx")])
        if self.filepath:
            self.update_info_var(f"File chosen: {self.filepath}")
            self.update_text_area(f"Chosen Excel File: {self.filepath}")
        else:
            self.update_info_var("No file chosen.")
            self.update_text_area("No file chosen.")

    def update_progress(self, value):
        self.progress["value"] = value
        self.root.update_idletasks()

    def update_text_area(self, message):
        self.text_area.insert(tk.END, message + "\n")
        self.text_area.see(tk.END)
        with open("log.txt", "a") as log_file:
            log_file.write(f"{time.ctime()}: {message}\n")

    def _safe_update_text_area(self, message):
        self.root.after(0, lambda: self.update_text_area(message))

    def update_info_var(self, message):
        self.root.after(0, self._safe_update_info_var, message)

    def _safe_update_info_var(self, message):
        self.info_var.set(message)

    # All send_messages and unsent_messages related code remains unchanged


    def set_webdriver_path(self):
        self.webdriver_path = filedialog.askopenfilename(title="Select WebDriver", filetypes=[("Executable files", "*.exe"), ("All files", "*.*")])
        if self.webdriver_path and os.path.isfile(self.webdriver_path):
            self.update_info_var(f"WebDriver set to: {self.webdriver_path}")
            self.update_text_area(f"WebDriver set to: {self.webdriver_path}")
        else:
            self.update_info_var("Invalid WebDriver path!")
            self.update_text_area("Invalid WebDriver path!")

    def login(self):
        if not self.webdriver_path:
            self.update_info_var("Please select WebDriver path first!")
            self.update_text_area("Please select WebDriver path first!")
            return

        chrome_options = Options()
        data_dir = os.path.join(os.getcwd(), 'whatsapp_session')
        chrome_options.add_argument(f"--user-data-dir={data_dir}")

        try:
            service = webdriver.chrome.service.Service(self.webdriver_path)
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            self.wait = WebDriverWait(self.driver, 10)
        except WebDriverException as e:
            self.update_info_var(f"Failed to launch WebDriver: {str(e)}")
            self.update_text_area(f"Error launching WebDriver: {str(e)}")
            return

        try:
            self.driver.get("https://web.whatsapp.com/")
            self.update_info_var("Status: Please login to WhatsApp...")
            self.update_text_area("Opened WhatsApp Web. Please login.")
        except Exception as e:
            self.update_info_var(f"Failed to open WhatsApp Web: {str(e)}")
            self.update_text_area(f"Error opening WhatsApp Web: {str(e)}")

    def choose_file(self):
        self.filepath = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel files", "*.xlsx")])
        if self.filepath:
            self.update_info_var(f"File chosen: {self.filepath}")
            self.update_text_area(f"Chosen Excel File: {self.filepath}")
        else:
            self.update_info_var("No file chosen.")
            self.update_text_area("No file chosen.")

    def send_messages(self):
        self.stop_thread = False

        if not self.driver:
            self.update_info_var("Please login first!")
            self.update_text_area("Please login first!")
            return
        if not self.filepath:
            self.update_info_var("Please choose an Excel file first!")
            self.update_text_area("Please choose an Excel file first!")
            return

        self.update_info_var("Status: Sending messages...")
        self.update_text_area("Started sending messages...")

        try:
            wb = openpyxl.load_workbook(self.filepath)
            sheet = wb.active

            # Workbooks for sent and unsent messages
            wb_sent = openpyxl.Workbook()
            sheet_sent = wb_sent.active
            sheet_sent.append(["Phone Number", "Message", "Status"])

            wb_unsent = openpyxl.Workbook()
            sheet_unsent = wb_unsent.active
            for row in sheet.iter_rows(values_only=True):
                sheet_unsent.append(row)

            unsent_row = 2
            total_rows = sheet.max_row - 1
            start_time = datetime.now()

            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
                if self.stop_thread:
                    self.update_text_area("Stopping the message sending process...")
                    break
                while self.pause_thread:
                    self.update_info_var("Status: Paused. Click Resume to continue.")
                    time.sleep(1)

                phone_number, message = row
                phone_number = str(phone_number)
                message = str(message)

                try:
                    self.driver.get(f"https://web.whatsapp.com/send?phone={phone_number}")
                    self.update_text_area(f"Attempting to send to {phone_number}...")

                    retries = 0
                    message_box = None
                    while retries < 3:
                        try:
                            if self.chosen_language.get() == "en":
                                message_box = self.wait.until(EC.presence_of_element_located((By.XPATH, "//div[@aria-placeholder='Type a message']")))
                            elif self.chosen_language.get() == "ar":
                                message_box = self.wait.until(EC.presence_of_element_located((By.XPATH, "//div[@aria-placeholder='اكتب رسالة']")))
                            break
                        except TimeoutException:
                            retries += 1
                            self.update_info_var(f"Retry {retries} for {phone_number}")
                            self.update_text_area(f"Retry {retries}: Failed to locate message box for {phone_number}")
                            continue

                    if not message_box:
                        self.update_text_area(f"Failed to find message box for {phone_number} after retries.")
                        unsent_row += 1
                        wb_unsent.save("unsent_messages.xlsx")  # Save after each unsent message
                        continue

                    if self.send_mode.get() == "message":
                        cleaned_message = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\xff]', '', message)
                        pyperclip.copy(cleaned_message)
                        message_box.send_keys(Keys.CONTROL, 'v')
                        message_box.send_keys(Keys.ENTER)
                    else:
                        if self.photo_path:
                            self.copy_image_to_clipboard(self.photo_path)
                            message_box.send_keys(Keys.CONTROL, 'v')
                            message_box.send_keys(Keys.ENTER)
                            try:
                                if self.chosen_language.get() == "en":
                                    message_box = self.wait.until(EC.presence_of_element_located((By.XPATH, "//div[@aria-placeholder='add caption']")))
                                elif self.chosen_language.get() == "ar":
                                    message_box = self.wait.until(EC.presence_of_element_located((By.XPATH, "//div[@aria-placeholder='إضافة شرح']")))
                            except TimeoutException:
                                    retries += 1
                                    continue
                            message_box.send_keys(Keys.ENTER)
                        else:
                            self.update_info_var(f"No photo chosen for {phone_number}")
                            self.update_text_area(f"No photo chosen for {phone_number}")
                            unsent_row += 1
                            sheet_unsent.append([phone_number, message])  # Add unsent message
                            wb_unsent.save("unsent_messages.xlsx")  # Save after each unsent message
                            continue

                    # Check message status
                    status = self.check_message_status()
                    self.update_text_area(f"Message/photo to {phone_number} {status}.")
                    
                    # Update sent and unsent logs accordingly
                    if status == "Not Sent":
                        unsent_row += 1
                    else:
                        sheet_sent.append([phone_number, message, status])  # Add sent message
                        sheet_unsent.delete_rows(unsent_row)


                    # Save both workbooks after each iteration
                    wb_sent.save("sent_messages.xlsx")
                    wb_unsent.save("unsent_messages.xlsx")

                except (NoSuchElementException, TimeoutException, WebDriverException) as e:
                    error_msg = f"Error for {phone_number}: {str(e)}"
                    self.update_info_var(f"Error for {phone_number}")
                    self.update_text_area(error_msg)
                    unsent_row += 1
                    wb_unsent.save("unsent_messages.xlsx")  # Save after each unsent message

                time.sleep(2)

                progress_value = (idx / total_rows) * 100
                self.update_progress(progress_value)

                elapsed_time = datetime.now() - start_time
                estimated_total_time = (elapsed_time / idx) * total_rows
                remaining_time = estimated_total_time - elapsed_time
                self.time_var.set(f"Estimated Time Remaining: {str(remaining_time).split('.')[0]}")

        except Exception as e:
            self.update_info_var(f"Error: {str(e)}")
            self.update_text_area(f"Error during message sending process: {str(e)}")
        finally:
            self.driver.quit()
            self.update_info_var("Status: Done sending messages!")
            self.update_text_area("Finished sending messages.")






    def check_message_status(self):
        icons = [
            ("//span[@data-icon='msg-time']", "Pending"),
            ("//span[@data-icon='msg-check']", "Sent"),
            ("//span[@data-icon='msg-dblcheck']", "Delivered"),
            ("//span[@data-icon='msg-dblcheck-ack']", "Read")
        ]

        retries = 0
        max_retries = 5
        wait_time = 2  # seconds

        # Retry until we either find a status or exhaust retries
        while retries < max_retries:
            for icon, state in icons:
                try:
                    # Wait for the icon to appear indicating message status
                    self.wait.until(EC.presence_of_element_located((By.XPATH, icon)))
                    self.update_text_area(f"Message status updated to: {state}")
                    
                    # If we reach this point, we've successfully found the status
                    return state

                except TimeoutException:
                    continue  # If the icon wasn't found, continue to the next one

            # Increment the retry counter and sleep for a while before the next retry
            retries += 1
            self.update_text_area(f"Retrying to check message status... (Attempt {retries}/{max_retries})")
            time.sleep(wait_time)

        # If we exit the loop without finding a status, assume it wasn't sent
        self.update_text_area("No final status found after retries. Returning 'Not Sent'.")
        return "Not Sent"

    def pause_messages(self):
        self.pause_thread = not self.pause_thread
        if self.pause_thread:
            self.update_info_var("Paused...")
        else:
            self.update_info_var("Resuming...")

    def stop_messages(self):
        self.stop_thread = True
        self.update_info_var("Stopping...")
        self.update_text_area("Attempting to stop the message sending process...")


if __name__ == "__main__":
    root = tk.Tk()
    app = WhatsAppAutomation(root)
    root.mainloop()

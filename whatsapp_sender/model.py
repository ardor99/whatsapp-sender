# whatsapp_sender/model.py

"""Model module for WhatsApp Automation application."""

import os
import re
import time

import pyperclip
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys

from whatsapp_sender.utils import copy_image_to_clipboard

class WhatsAppModelError(Exception):
    """Custom exception class for WhatsAppModel errors."""

class WhatsAppModel:
    """Model class that handles data and interactions with WhatsApp Web."""

    # pylint: disable=too-many-instance-attributes

    def __init__(self):
        """Initialize the model with default values."""
        self.driver = None
        self.wait = None
        self.webdriver_path = ''
        self.photo_path = ''
        self.messages = []
        self.sent_messages = []
        self.unsent_messages = []
        self.chosen_language = 'en'

    def set_webdriver_path(self, path):
        """Set the path to the WebDriver executable."""
        self.webdriver_path = path

    def set_language(self, lang):
        """Set the language for the WhatsApp Web interface."""
        self.chosen_language = lang

    def login(self):
        """Initialize the WebDriver and navigate to WhatsApp Web."""
        if not self.webdriver_path:
            raise ValueError("WebDriver path is not set.")

        chrome_options = Options()
        data_dir = os.path.join(os.getcwd(), 'whatsapp_session')
        chrome_options.add_argument(f"--user-data-dir={data_dir}")

        try:
            service = webdriver.chrome.service.Service(self.webdriver_path)
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            self.wait = WebDriverWait(self.driver, 10)
            self.driver.get("https://web.whatsapp.com/")
        except WebDriverException as e:
            raise WhatsAppModelError(f"Failed to launch WebDriver: {str(e)}") from e

    def load_excel_file(self, filepath):
        """Load messages from the specified Excel file."""
        try:
            wb = openpyxl.load_workbook(filepath)
            sheet = wb.active
            self.messages = [
                (str(row[0].value), str(row[1].value))
                for row in sheet.iter_rows(min_row=2)
            ]
        except Exception as e:
            raise WhatsAppModelError(f"Failed to load Excel file: {str(e)}") from e

    def send_message(self, phone_number, message):
        """Send a text message to the specified phone number."""
        try:
            self.driver.get(f"https://web.whatsapp.com/send?phone={phone_number}")
            retries = 0
            message_box = None
            while retries < 3:
                try:
                    if self.chosen_language == "en":
                        message_box = self.wait.until(EC.presence_of_element_located(
                            (By.XPATH, "//div[@aria-placeholder='Type a message']")))
                    elif self.chosen_language == "ar":
                        message_box = self.wait.until(EC.presence_of_element_located(
                            (By.XPATH, "//div[@aria-placeholder='اكتب رسالة']")))
                    break
                except TimeoutException:
                    retries += 1
                    time.sleep(1)
                    continue

            if not message_box:
                return "Failed to locate message box."

            cleaned_message = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', message)
            pyperclip.copy(cleaned_message)
            message_box.send_keys(Keys.CONTROL, 'v')
            message_box.send_keys(Keys.ENTER)

            status = self.check_message_status()
            self.sent_messages.append((phone_number, message, status))
            return status
        except (TimeoutException, WebDriverException) as e:
            self.unsent_messages.append((phone_number, message))
            return f"Error: {str(e)}"

    def send_photo(self, phone_number, message):
        """Send a photo with an optional caption to the specified phone number."""
        try:
            self.driver.get(f"https://web.whatsapp.com/send?phone={phone_number}")
            retries = 0
            message_box = None
            while retries < 3:
                try:
                    if self.chosen_language == "en":
                        message_box = self.wait.until(EC.presence_of_element_located(
                            (By.XPATH, "//div[@aria-placeholder='Type a message']")))
                    elif self.chosen_language == "ar":
                        message_box = self.wait.until(EC.presence_of_element_located(
                            (By.XPATH, "//div[@aria-placeholder='اكتب رسالة']")))
                    break
                except TimeoutException:
                    retries += 1
                    time.sleep(1)
                    continue

            if not message_box:
                return "Failed to locate message box."

            copy_image_to_clipboard(self.photo_path)
            message_box.send_keys(Keys.CONTROL, 'v')
            time.sleep(1)
            caption_box = self.wait.until(EC.presence_of_element_located(
                (By.XPATH, "//div[@data-testid='media-caption-text-input']")))
            pyperclip.copy(message)
            caption_box.send_keys(Keys.CONTROL, 'v')
            send_button = self.wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//span[@data-testid='send']")))
            send_button.click()

            status = self.check_message_status()
            self.sent_messages.append((phone_number, message, status))
            return status
        except (TimeoutException, WebDriverException) as e:
            self.unsent_messages.append((phone_number, message))
            return f"Error: {str(e)}"

    def check_message_status(self):
        """Check the status of the last sent message."""
        icons = [
            ("//span[@data-icon='msg-time']", "Pending"),
            ("//span[@data-icon='msg-check']", "Sent"),
            ("//span[@data-icon='msg-dblcheck']", "Delivered"),
            ("//span[@data-icon='msg-dblcheck-ack']", "Read")
        ]

        retries = 0
        max_retries = 5
        wait_time = 2

        while retries < max_retries:
            for icon, state in icons:
                try:
                    self.wait.until(EC.presence_of_element_located((By.XPATH, icon)))
                    return state
                except TimeoutException:
                    continue
            retries += 1
            time.sleep(wait_time)

        return "Not Sent"

    def save_sent_messages(self):
        """Save the list of sent messages to an Excel file."""
        wb_sent = openpyxl.Workbook()
        sheet_sent = wb_sent.active
        sheet_sent.append(["Phone Number", "Message", "Status"])
        for record in self.sent_messages:
            sheet_sent.append(record)
        wb_sent.save("sent_messages.xlsx")

    def save_unsent_messages(self):
        """Save the list of unsent messages to an Excel file."""
        wb_unsent = openpyxl.Workbook()
        sheet_unsent = wb_unsent.active
        sheet_unsent.append(["Phone Number", "Message"])
        for record in self.unsent_messages:
            sheet_unsent.append(record)
        wb_unsent.save("unsent_messages.xlsx")

    def cleanup(self):
        """Clean up resources and save message logs."""
        self.save_sent_messages()
        self.save_unsent_messages()
        if self.driver:
            self.driver.quit()
